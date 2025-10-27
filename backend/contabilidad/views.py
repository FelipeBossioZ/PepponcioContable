#contabilidad/views.py

from rest_framework import viewsets, filters, views
from rest_framework.response import Response
#from rest_framework.permissions import AllowAny   # 👈 añade esto
from .models import Cuenta, AsientoContable, MovimientoContable
from .serializers import CuentaSerializer, AsientoContableSerializer, MovimientoContableSerializer
from django.utils import timezone
from datetime import date
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import PeriodoContable
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from django.db.models import Sum
from decimal import Decimal
from datetime import datetime
from rest_framework.decorators import api_view, permission_classes

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def periodo_view(request):
    anio = int(request.query_params.get('anio'))
    p = PeriodoContable.ensure(anio)
    return Response({
        "anio": p.anio,
        "estado": p.estado,
        "ajustes": {
            "inicio": p.ajustes_inicio.isoformat(),
            "fin":    p.ajustes_fin.isoformat(),
            "requiere_pins": p.requiere_pins_en_ajustes,
        },
        "mes13": {
            "habilitado": p.habilitar_mes13,
            "etiqueta": "Mes 13 (Ajustes de cierre)"
        }
    })

def _parse_date(val):
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):  # ISO y día/mes/año
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None



class CuentaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para visualizar las Cuentas Contables.

    Proporciona las acciones `list` y `retrieve` (solo lectura).
    Permite filtrar por el nombre de la cuenta.
    """
    
    queryset = Cuenta.objects.all().order_by("codigo")
    serializer_class = CuentaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["codigo", "nombre"]
    pagination_class = None

class AsientoContableViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión de Asientos Contables.

    Permite crear, listar y ver el detalle de los asientos.
    La actualización y eliminación se deshabilitan por seguridad contable,
    ya que los asientos no deben modificarse (se deben crear asientos de ajuste).
    """
    permission_classes = [IsAuthenticated]
    queryset = AsientoContable.objects.prefetch_related("movimientos").all()
    serializer_class = AsientoContableSerializer
    http_method_names = ['get', 'post', 'head', 'options']  # Deshabilitar PUT, PATCH, DELETE

    def get_queryset(self):
        """
        Opcionalmente filtra los asientos por un rango de fechas (?fecha_inicio, ?fecha_fin).
        """
        qs = super().get_queryset()
        fi = _parse_date(self.request.query_params.get('fecha_inicio'))
        ff = _parse_date(self.request.query_params.get('fecha_fin'))
        if fi:
            qs = qs.filter(fecha__gte=fi)
        if ff:
            qs = qs.filter(fecha__lte=ff)
        return qs

    @action(detail=True, methods=["post"], url_path="anular")
    def anular(self, request, pk=None):
        asiento = self.get_object()
        if asiento.estado == "anulado":
            return Response({"detail": "El asiento ya está anulado."}, status=400)

        hoy = timezone.localdate()
        # Tomamos el año fiscal del asiento si existe, si no, el de la fecha
        anio_fiscal = asiento.fiscal_year or asiento.fecha.year
        periodo = PeriodoContable.ensure(anio_fiscal)

        # Estado del periodo
        if periodo.estado == "cerrado":
            return Response(
                {"detail": f"El periodo {periodo.anio} está CERRADO. Anulación prohibida."},
                status=403
            )

        # Regla de ventanas:
        #  - Hasta 31/12 del año fiscal: anula sin PIN
        #  - Durante la ventana de ajustes (enero–marzo del año siguiente): requiere PIN si está configurado
        #  - Fuera de eso: prohibido
        limite_libre = date(anio_fiscal, 12, 31)
        requiere_pins = False

        if hoy <= limite_libre:
            requiere_pins = False
        elif periodo.in_ajustes(hoy) and hoy.year == anio_fiscal + 1:
            requiere_pins = bool(periodo.requiere_pins_en_ajustes)
        else:
            return Response({"detail": "Fuera de la ventana permitida para anulación."}, status=403)

        if requiere_pins:
            from django.conf import settings
            contador_pin = request.data.get("contador_pin")
            gerente_pin  = request.data.get("gerente_pin")
            ok_cont = bool(contador_pin) and (contador_pin == getattr(settings, "CONTADOR_PIN", None))
            ok_ger  = bool(gerente_pin)  and (gerente_pin  == getattr(settings, "GERENTE_PIN", None))
            if not (ok_cont and ok_ger):
                return Response({"detail": "Se requieren PINs válidos de Contador y Gerente."}, status=403)

        motivo = request.data.get("motivo", "")

        # Crear asiento de ajuste (inverso) y marcar anulado
        ajuste = AsientoContable.objects.create(
            fecha=hoy,
            concepto=f"AJUSTE POR ANULACIÓN del asiento #{asiento.id}",
            tercero=asiento.tercero,
            descripcion_adicional=f"Motivo: {motivo}",
        )
        for m in asiento.movimientos.all():
            MovimientoContable.objects.create(
                asiento=ajuste,
                cuenta=m.cuenta,
                debito=m.credito,
                credito=m.debito,
            )

        asiento.estado = "anulado"
        asiento.anulado_por = request.user
        asiento.anulado_en = timezone.now()
        asiento.anulacion_motivo = motivo
        asiento.ajusta_a = ajuste
        asiento.save()

        return Response({"detail": "Asiento anulado y ajuste generado", "ajuste_id": ajuste.id}, status=200)

class LibroDiarioView(views.APIView):
    """
    Vista para generar el reporte de Libro Diario.
    Devuelve todos los movimientos contables ordenados por fecha.
    """
    def get(self, request):
        fi = _parse_date(request.query_params.get('fecha_inicio'))
        ff = _parse_date(request.query_params.get('fecha_fin'))

        movimientos = (
            MovimientoContable.objects
            .select_related('asiento', 'cuenta')
            .order_by('asiento__fecha', 'asiento__id')
        )
        if fi:
            movimientos = movimientos.filter(asiento__fecha__gte=fi)
        if ff:
            movimientos = movimientos.filter(asiento__fecha__lte=ff)

        data = []
        for m in movimientos:
            data.append({
                'fecha': m.asiento.fecha,
                'asiento_id': m.asiento.id,
                'tercero': getattr(m.asiento.tercero, 'nombre_razon_social', None),
                'codigo_cuenta': m.cuenta.codigo,
                'nombre_cuenta': m.cuenta.nombre,
                'concepto': m.asiento.concepto,
                'debito': m.debito,
                'credito': m.credito,
            })
        return Response(data, status=200)


class BalancePruebasView(views.APIView):
    def get(self, request):
        fi = _parse_date(request.query_params.get('fecha_inicio'))
        ff = _parse_date(request.query_params.get('fecha_fin'))

        cuentas = Cuenta.objects.all().order_by('codigo')

        # Movimientos del período
        movs = MovimientoContable.objects.all()
        if fi:
            movs = movs.filter(asiento__fecha__gte=fi)
        if ff:
            movs = movs.filter(asiento__fecha__lte=ff)

        periodo = movs.values('cuenta__codigo','cuenta__nombre') \
                      .annotate(total_debito=Sum('debito'), total_credito=Sum('credito')) \
                      .order_by('cuenta__codigo')
        periodo_dict = {
            x['cuenta__codigo']: {
                'nombre': x['cuenta__nombre'],
                'deb': x['total_debito'] or Decimal('0'),
                'cre': x['total_credito'] or Decimal('0'),
            } for x in periodo
        }

        # Saldos anteriores (saldo inicial)
        prev_dict = {}
        if fi:
            movs_prev = MovimientoContable.objects.filter(asiento__fecha__lt=fi)
            prev = movs_prev.values('cuenta__codigo').annotate(deb=Sum('debito'), cre=Sum('credito'))
            prev_dict = {
                x['cuenta__codigo']: (x['deb'] or Decimal('0')) - (x['cre'] or Decimal('0'))
                for x in prev
            }

        reporte = []
        total_ini = total_debitos = total_creditos = total_fin = Decimal('0')

        for cta in cuentas:
            ini = prev_dict.get(cta.codigo, Decimal('0'))
            deb = periodo_dict.get(cta.codigo, {}).get('deb', Decimal('0'))
            cre = periodo_dict.get(cta.codigo, {}).get('cre', Decimal('0'))
            if ini == 0 and deb == 0 and cre == 0:
                continue
            fin = ini + deb - cre
            reporte.append({
                'codigo_cuenta': cta.codigo,
                'nombre_cuenta': cta.nombre,
                'saldo_inicial': ini,
                'total_debito': deb,
                'total_credito': cre,
                'saldo_final': fin,
            })
            total_ini       += ini
            total_debitos   += deb
            total_creditos  += cre
            total_fin       += fin

        formato = request.query_params.get("formato")
        if formato == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance de Prueba"

            ws["A1"] = "NOMBRE DE LA EMPRESA + NIT"
            ws["A2"] = f"Balance de Prueba   {ff or fi or ''}"
            ws["A1"].font = ws["A2"].font = Font(bold=True)

            headers = ["Cuenta","Nombre Cuenta contable","Saldo inicial","Débitos","Créditos","Saldo final"]
            ws.append(headers)
            for cell in ws[3]:
                cell.font = Font(bold=True)

            for row in reporte:
                ws.append([
                    row["codigo_cuenta"],
                    row["nombre_cuenta"],
                    float(row["saldo_inicial"]),
                    float(row["total_debito"]),
                    float(row["total_credito"]),
                    float(row["saldo_final"]),
                ])

            ws.append(["","TOTAL", float(total_ini), float(total_debitos), float(total_creditos), float(total_fin)])
            widths = [12, 40, 16, 14, 14, 16]
            for i,w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
                for cell in ws[get_column_letter(i)]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="balance_prueba.xlsx"'
            wb.save(resp)
            return resp

        return Response({
            'detalle': reporte,
            'sumas_iguales': {
                'total_inicial': total_ini,
                'total_debitos': total_debitos,
                'total_creditos': total_creditos,
                'total_final': total_fin,
            }
        }, status=200)


class LibroMayorView(views.APIView):
    """
    Reporte de Libro Mayor para una cuenta (por código).
    GET /api/contabilidad/libro-mayor/<codigo_cuenta>/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
    """
    def get(self, request, codigo_cuenta):
        try:
            cuenta = Cuenta.objects.get(pk=codigo_cuenta)   # tu PK de cuenta es 'codigo'
        except Cuenta.DoesNotExist:
            return Response({"error": "La cuenta especificada no existe."}, status=404)

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin    = request.query_params.get('fecha_fin')

        # Saldo inicial
        mov_prev = MovimientoContable.objects.filter(cuenta=cuenta)
        if fecha_inicio:
            mov_prev = mov_prev.filter(asiento__fecha__lt=fecha_inicio)
        agg_prev = mov_prev.aggregate(total_debito=Sum('debito'), total_credito=Sum('credito'))
        deb_prev = agg_prev['total_debito']  or Decimal('0')
        cre_prev = agg_prev['total_credito'] or Decimal('0')
        saldo_inicial = deb_prev - cre_prev

        # Movimientos del período
        qs = MovimientoContable.objects.filter(cuenta=cuenta)
        if fecha_inicio:
            qs = qs.filter(asiento__fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(asiento__fecha__lte=fecha_fin)
        qs = qs.select_related('asiento', 'asiento__tercero').order_by('asiento__fecha','asiento__id')

        detalle = []
        saldo = saldo_inicial
        for m in qs:
            saldo += m.debito - m.credito
            detalle.append({
                'fecha': m.asiento.fecha,
                'asiento_id': m.asiento.id,
                'tercero': getattr(m.asiento.tercero, 'nombre_razon_social', None),
                'concepto': m.asiento.concepto,
                'debito': m.debito,
                'credito': m.credito,
                'saldo': saldo,
            })

        return Response({
            'cuenta': {'codigo': cuenta.codigo, 'nombre': cuenta.nombre},
            'fecha_inicio_reporte': fecha_inicio,
            'fecha_fin_reporte': fecha_fin,
            'saldo_inicial': saldo_inicial,
            'movimientos': detalle,
            'saldo_final': saldo,
        }, status=200)


class EstadoResultadosView(views.APIView):
    """
    Vista para generar el Estado de Resultados a una fecha de corte.
    """
    

    def get(self, request):
        from django.db.models import Sum
        from decimal import Decimal

        fecha_fin = request.query_params.get('fecha_fin')
        if not fecha_fin:
            return Response({"error": "Debe proporcionar una 'fecha_fin' en los parámetros."}, status=400)

        # 1. Ingresos (Clase 4)
        ingresos_agg = MovimientoContable.objects.filter(
            asiento__fecha__lte=fecha_fin,
            cuenta__codigo__startswith='4'
        ).aggregate(total_credito=Sum('credito'), total_debito=Sum('debito'))
        total_ingresos = (ingresos_agg['total_credito'] or 0) - (ingresos_agg['total_debito'] or 0)

        # 2. Costos de Ventas (Clase 6)
        costos_agg = MovimientoContable.objects.filter(
            asiento__fecha__lte=fecha_fin,
            cuenta__codigo__startswith='6'
        ).aggregate(total_debito=Sum('debito'), total_credito=Sum('credito'))
        total_costos = (costos_agg['total_debito'] or 0) - (costos_agg['total_credito'] or 0)

        utilidad_bruta = total_ingresos - total_costos

        # 3. Gastos Operacionales (Clase 5)
        gastos_agg = MovimientoContable.objects.filter(
            asiento__fecha__lte=fecha_fin,
            cuenta__codigo__startswith='5'
        ).aggregate(total_debito=Sum('debito'), total_credito=Sum('credito'))
        total_gastos = (gastos_agg['total_debito'] or 0) - (gastos_agg['total_credito'] or 0)

        utilidad_operacional = utilidad_bruta - total_gastos

        # Por ahora, un cálculo simplificado. Se pueden añadir más niveles (no operacionales, impuestos, etc.)
        utilidad_neta_antes_impuestos = utilidad_operacional

        return Response({
            'fecha_corte': fecha_fin,
            'ingresos_operacionales': total_ingresos,
            'costo_de_ventas': total_costos,
            'utilidad_bruta': utilidad_bruta,
            'gastos_operacionales': total_gastos,
            'utilidad_operacional': utilidad_operacional,
            'utilidad_neta_antes_de_impuestos': utilidad_neta_antes_impuestos
        })

class BalanceGeneralView(views.APIView):
    """
    Vista para generar el Balance General (Estado de Situación Financiera).
    """
    

    def get(self, request):
        from django.db.models import Sum
        from decimal import Decimal

        fecha_fin = request.query_params.get('fecha_fin')
        if not fecha_fin:
            return Response({"error": "Debe proporcionar una 'fecha_fin' en los parámetros."}, status=400)

        # Función auxiliar para calcular el saldo de una clase de cuenta
        def calcular_saldo_clase(clase):
            movimientos = MovimientoContable.objects.filter(
                asiento__fecha__lte=fecha_fin,
                cuenta__codigo__startswith=str(clase)
            )
            saldo = movimientos.aggregate(
                total_debito=Sum('debito', default=Decimal(0)),
                total_credito=Sum('credito', default=Decimal(0))
            )
            return saldo['total_debito'] - saldo['total_credito']

        # Calcular saldos para Activo (1), Pasivo (2), y Patrimonio (3)
        total_activos = calcular_saldo_clase(1)
        total_pasivos = calcular_saldo_clase(2) * -1 # Se multiplica por -1 porque los pasivos son de naturaleza crédito
        total_patrimonio_inicial = calcular_saldo_clase(3) * -1

        # Calcular la utilidad del ejercicio (Ingresos - Gastos - Costos)
        utilidad_ingresos = calcular_saldo_clase(4) * -1
        utilidad_gastos = calcular_saldo_clase(5)
        utilidad_costos = calcular_saldo_clase(6)
        utilidad_del_ejercicio = utilidad_ingresos - utilidad_gastos - utilidad_costos

        total_patrimonio = total_patrimonio_inicial + utilidad_del_ejercicio

        # Verificación de la ecuación contable
        ecuacion_ok = total_activos == (total_pasivos + total_patrimonio)

        return Response({
            'fecha_corte': fecha_fin,
            'activos': {
                'total': total_activos,
            },
            'pasivos_y_patrimonio': {
                'total_pasivos': total_pasivos,
                'total_patrimonio': total_patrimonio,
                'total_pasivo_y_patrimonio': total_pasivos + total_patrimonio,
            },
            'verificacion_ecuacion_contable': {
                'ecuacion': 'Activo = Pasivo + Patrimonio',
                'balance_correcto': ecuacion_ok
            }
        })

class MediosMagneticosView(views.APIView):
    """
    Vista para generar reportes de Medios Magnéticos (Exógena) para la DIAN.
    """
    

    def get(self, request, formato):
        year = request.query_params.get('year')
        if not year:
            return Response({"error": "Debe proporcionar el parámetro 'year'."}, status=400)

        if formato == '1001':
            return self.formato_1001(year)

        # Aquí se añadirían las llamadas a otros formatos (1003, 1005, etc.)

        return Response({"error": f"El formato '{formato}' no es soportado aún."}, status=404)

    def formato_1001(self, year):
        """
        Genera los datos para el Formato 1001 v10: Pagos o abonos en cuenta.
        Conceptos de ejemplo: 5002 (Salarios), 5016 (Honorarios).
        """
        from django.db.models import Sum, Value, CharField
        from django.db.models.functions import Coalesce

        # Filtra movimientos de cuentas de Gasto (clase 5) para el año especificado.
        pagos = MovimientoContable.objects.filter(
            asiento__fecha__year=year,
            cuenta__codigo__startswith='5'
        ).values(
            'asiento__tercero__tipo_documento',
            'asiento__tercero__numero_documento',
            'asiento__tercero__nombre_razon_social'
        ).annotate(
            # Agregamos el concepto basado en la cuenta. Lógica de ejemplo.
            concepto_pago=Value('5016', output_field=CharField()), # Asumimos honorarios por defecto
            valor_pago=Sum('debito')
        ).filter(valor_pago__gt=0) # Solo pagos, no notas crédito

        # Aquí iría una lógica más compleja para mapear cuentas a conceptos DIAN.
        # Por ahora, es una implementación de ejemplo.

        data = list(pagos)

        return Response({
            'formato': '1001',
            'version': '10',
            'año': year,
            'registros': data
        })

        